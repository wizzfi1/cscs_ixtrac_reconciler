import os
import pandas as pd
from openpyxl import load_workbook

from core.loader import load_excel
from core.normalizer import normalize_name, first_two_names
from core.matcher import build_cscs_index, build_cscs_index_2name
from core.validator import validate_membercode
from core.mapping import load_mappings, validate_mapping, resolve_columns
from config.rules import (
    STATUS_CONFIRMED,
    STATUS_CONFIRMED_2NAME,
    STATUS_AMBIGUOUS,
    STATUS_NOT_FOUND,
    STATUS_PRIORITY,
)


def run_reconciliation(file_path: str, mapping_name: str):
    """
    Main reconciliation entry point.
    Uses saved mapping to reconcile CSCS → IX TRAC safely.
    """

    # =========================
    # Setup
    # =========================
    os.makedirs("output", exist_ok=True)
    output_path = "output/IXTRAC_RECONCILED.xlsx"

    mappings = load_mappings()
    if mapping_name not in mappings:
        raise ValueError(f"Unknown mapping: {mapping_name}")

    mapping = mappings[mapping_name]

    # Sheets
    cscs_sheet = mapping.get("cscs_sheet", "CSCS")
    ixtrac_sheet = mapping.get("ixtrac_sheet") or mapping.get("sheet")

    # Columns
    cscs_name_col = mapping.get("cscs_name", "NAME")

    # =========================
    # Load CSCS (SOURCE)
    # =========================
    cscs = load_excel(file_path, cscs_sheet)

    if cscs_name_col not in cscs.columns:
        raise ValueError(
            f"CSCS name column '{cscs_name_col}' not found in sheet '{cscs_sheet}'."
        )

    cscs["NORM_NAME"] = cscs[cscs_name_col].apply(normalize_name)
    cscs["FIRST2"] = cscs[cscs_name_col].apply(first_two_names)

    exact_index = build_cscs_index(cscs)
    two_name_index = build_cscs_index_2name(cscs)

    # =========================
    # Load IX TRAC (TARGET)
    # =========================
    wb = load_workbook(file_path)
    if ixtrac_sheet not in wb.sheetnames:
        raise ValueError(f"IX TRAC sheet '{ixtrac_sheet}' not found.")

    sheet = wb[ixtrac_sheet]

    # Validate mapping inputs
    validate_mapping(sheet, mapping)
    cols = resolve_columns(sheet, mapping)

    # =========================
    # PHASE 1 — ENRICH IX TRAC
    # =========================
    for r in range(2, sheet.max_row + 1):
        name = sheet.cell(r, cols["name"]).value
        chn = sheet.cell(r, cols["chn"]).value

        membercode = ""
        status = STATUS_NOT_FOUND

        if name and chn:
            norm = normalize_name(name)
            matches = exact_index.get((norm, chn), [])

            valid = []
            terminal_status = None

            # --- Exact match ---
            for m in matches:
                ok, reason = validate_membercode(m["MEMBERCODE"])
                if ok:
                    valid.append(m)
                elif terminal_status is None and reason:
                    terminal_status = reason

            if len(valid) == 1:
                membercode = valid[0]["MEMBERCODE"]
                status = STATUS_CONFIRMED

            elif len(valid) > 1:
                status = STATUS_AMBIGUOUS

            elif terminal_status:
                status = terminal_status

            else:
                # --- First two names + CHN ---
                first2 = first_two_names(name)
                matches = two_name_index.get((first2, chn), [])

                valid = []
                terminal_status = None

                for m in matches:
                    ok, reason = validate_membercode(m["MEMBERCODE"])
                    if ok:
                        valid.append(m)
                    elif terminal_status is None and reason:
                        terminal_status = reason

                if len(valid) == 1:
                    membercode = valid[0]["MEMBERCODE"]
                    status = STATUS_CONFIRMED_2NAME

                elif len(valid) > 1:
                    status = STATUS_AMBIGUOUS

                elif terminal_status:
                    status = terminal_status

        sheet.cell(r, cols["membercode"]).value = membercode
        sheet.cell(r, cols["status"]).value = status

    # Save enriched IX TRAC (no sorting)
    wb.save(output_path)

    # =========================
    # PHASE 2 — REVIEW + SUMMARY
    # =========================
    ix_df = pd.read_excel(output_path, sheet_name=ixtrac_sheet)
    status_col = mapping["status_out"]

    # --- REVIEW (sorted copy) ---
    review_df = ix_df.copy()
    review_df["__rank"] = review_df[status_col].map(STATUS_PRIORITY)
    review_df = review_df.sort_values("__rank").drop(columns="__rank")

    # --- SUMMARY ---
    summary_df = (
        review_df[status_col]
        .value_counts()
        .reset_index()
        .rename(columns={
            "index": "MATCH_STATUS",
            status_col: "COUNT"
        })
    )

    summary_df.loc[len(summary_df)] = {
        "MATCH_STATUS": "TOTAL_ROWS",
        "COUNT": len(review_df)
    }

    # Append sheets safely
    with pd.ExcelWriter(
        output_path,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="replace"
    ) as writer:
        review_df.to_excel(writer, sheet_name="IX_TRAC_REVIEW", index=False)
        summary_df.to_excel(writer, sheet_name="RECONCILIATION_SUMMARY", index=False)

    print("✔ Reconciliation complete")
    print(f"✔ Output written to {output_path}")
